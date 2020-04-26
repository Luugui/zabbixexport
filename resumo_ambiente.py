#!/usr/bin/python3

from pyzabbix import ZabbixAPI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pyfiglet import Figlet
from tqdm import tqdm
import datetime, argparse


parser = argparse.ArgumentParser(description="Extracao de dados do Zabbix")
parser.add_argument('-u', '--user', required=True, help="Usuario do Zabbix (Admin)")
parser.add_argument('-p', '--password', required=True, help="Senha do Zabbix (zabbix)")
parser.add_argument('-s', '--server', required=True, help="Endereco Zabbix (http://localhost/zabbix)")
parser.add_argument('-g', '--group', action="append", help="ID de grupos de Host")
parser.add_argument('-n', '--name', default="Zabbix", help="Nome do relatorio gerado ex: Name_DATAHOJE.xlsx")
args = vars(parser.parse_args())

f = Figlet(font='slant')
print(f.renderText("Zabbix Export All Hosts"))


# CONECTANDO AO ZABBIX TAESA
zapi = ZabbixAPI(args['server'])
zapi.login(args['user'],args['password'])
print("--> Conectado com sucesso!\n")

if args['group']:
    for g in zapi.hostgroup.get(output="extend", groupids=args['group']):
     print("--> Grupo selecionado: "+g['name'])

# CRIANDO PLANILHA
wb = Workbook()
sheet = wb.active
sheet.title = "RESUMO"
sheet['A1'] = "GRUPO"
sheet['A1'].font = Font(sz=12, bold=True)
sheet['B1'] = "HOSTS"
sheet['B1'].font = Font(sz=12, bold=True)

row=3

# SEVERIDADE
SEV = {
    "0": "Not classified",
    "1": "Information",
    "2": "Warning",
    "3": "Average",
    "4": "High",
    "5": "Disaster"
}

colw = 0

# LISTANDO GRUPOS
print("\n--> Listando grupos e quantidade de hosts\n")
for g in tqdm(zapi.hostgroup.get(output="extend", groupids=args['group'])):
    if "Templates" not in g['name']:
        if len(g['name']) >= 16:
            GRUPO_NAME = g['name'].split(".")
            NAME = "#'{}'!A1".format(GRUPO_NAME[-1])
        else:
            NAME = "#'{}'!A1".format(g['name'])
        sheet.cell(row=row, column=1).value = '=HYPERLINK("{0}","{1}")'.format(NAME,g['name'])

        if len(g['name']) > colw:
            colw = len(g['name'])
        h = zapi.host.get(output="extend",groupids=g['groupid'])
        sheet.cell(row=row, column=2).value = len(h)
        #print("G: ",g['name']," N:",len(h))
        row+=1
sheet.column_dimensions['A'].width = colw+3


print("\n--> Gerando relatorio!\n")
for g in tqdm(zapi.hostgroup.get(output="extend", groupids=args['group'])):
    colA = 0
    colB = 0
    colC = 0
    if "Template" not in g['name']:
        GRUPO = g['name']
        if "/" in GRUPO:
            GRUPO = GRUPO.replace("/",".")
        #print('g: ',GRUPO)
        if len(GRUPO) >= 16:
            GRUPO_NAME = GRUPO.split(".")
            GRUPO = GRUPO_NAME[-1]
        wb.create_sheet(GRUPO)
        S_GRUPO = wb[GRUPO]
        S_GRUPO['A1'] = "HOST"
        S_GRUPO['A1'].font = Font(sz=12, bold=True)
        S_GRUPO['B1'] = "IP"
        S_GRUPO['B1'].font = Font(sz=12, bold=True)
        S_GRUPO['C1'] = "ITEM"
        S_GRUPO['C1'].font = Font(sz=12, bold=True)
        S_GRUPO['D1'] = '=HYPERLINK("#RESUMO!A1","VOLTAR")'
        S_GRUPO['D1'].font = Font(sz=12, bold=True)
        row=2
        for h in zapi.host.get(output="extend", groupids=g["groupid"]):
            for i in zapi.hostinterface.get(output="extend",hostids=h['hostid']):
                for item in zapi.item.get(output="extend",hostids=h['hostid']):
                        ativos = {"HOST":h['host'], "IP":i['ip'], "ITEM":item['name']}
                        #print('id: ',row," ",ativos)
                        S_GRUPO.cell(row=row, column=1).value = ativos['HOST']
                        if len(ativos['HOST']) > colA:
                            colA = len(ativos['HOST'])
                        S_GRUPO.cell(row=row, column=2).value = ativos['IP']
                        if len(ativos['IP']) > colB:
                            colB = len(ativos['IP'])
                        S_GRUPO.cell(row=row, column=3).value = ativos['ITEM']
                        if len(ativos['ITEM']) > colC:
                            colC = len(ativos['ITEM'])
                        row+=1
        dm = 'A1:C{}'.format(row)
        S_GRUPO.auto_filter.ref = dm
        S_GRUPO.freeze_panes = 'A2'
        S_GRUPO.column_dimensions['A'].width = colA+3
        S_GRUPO.column_dimensions['B'].width = colB
        S_GRUPO.column_dimensions['C'].width = colC

DATE = datetime.date.today()
NOME = args['name']+'_'+str(DATE)+'.xlsx'

wb.save(NOME)
print("\n--> Relatorio {0} Gerado!".format(NOME))
zapi.user.logout()
